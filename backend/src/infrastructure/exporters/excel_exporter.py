from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.worksheet.worksheet import Worksheet

from src.presentation.schemas.inward_report import InwardReportResponse
from src.presentation.schemas.sales_report import SalesReportResponse


def _write_filter_row(ws: Worksheet, filters: dict, col_count: int) -> None:
    if not filters:
        summary = "Filters: All"
    else:
        parts = [f"{key}: {value}" for key, value in filters.items() if value is not None]
        summary = "Filters: " + ", ".join(parts) if parts else "Filters: All"

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=col_count)
    ws.cell(1, 1, summary)


def _write_headers(ws: Worksheet, row: int, headers: list[str]) -> None:
    bold = Font(bold=True)
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row, col, header)
        cell.font = bold


class ExcelExporter:
    def export_sales_report(self, data: SalesReportResponse, filters: dict) -> BytesIO:
        wb = Workbook()

        ws_con = wb.active
        ws_con.title = "Consolidation"
        con_headers = ["Design", "Size", "Grade", "Total Nos"]
        _write_filter_row(ws_con, filters, len(con_headers))
        _write_headers(ws_con, 2, con_headers)
        for i, row in enumerate(data.consolidation, 3):
            ws_con.cell(i, 1, row.design_name)
            ws_con.cell(i, 2, row.size)
            ws_con.cell(i, 3, row.grade_code)
            ws_con.cell(i, 4, row.total_nos)

        ws_txn = wb.create_sheet("Transactions")
        txn_headers = ["Date", "Dealer", "Place", "Design", "Size", "Grade", "Nos"]
        _write_filter_row(ws_txn, filters, len(txn_headers))
        _write_headers(ws_txn, 2, txn_headers)
        for i, row in enumerate(data.transactions, 3):
            ws_txn.cell(i, 1, str(row.sales_date))
            ws_txn.cell(i, 2, row.dealer_name)
            ws_txn.cell(i, 3, row.place)
            ws_txn.cell(i, 4, row.design_name)
            ws_txn.cell(i, 5, row.size)
            ws_txn.cell(i, 6, row.grade_code)
            ws_txn.cell(i, 7, row.nos)

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    def export_inward_report(self, data: InwardReportResponse, filters: dict) -> BytesIO:
        wb = Workbook()

        ws_con = wb.active
        ws_con.title = "Consolidation"
        con_headers = ["Design", "Size", "Grade", "Total Nos"]
        _write_filter_row(ws_con, filters, len(con_headers))
        _write_headers(ws_con, 2, con_headers)
        for i, row in enumerate(data.consolidation, 3):
            ws_con.cell(i, 1, row.design_name)
            ws_con.cell(i, 2, row.size)
            ws_con.cell(i, 3, row.grade_code)
            ws_con.cell(i, 4, row.total_nos)

        ws_txn = wb.create_sheet("Transactions")
        txn_headers = ["Date", "Supplier", "Place", "Design", "Size", "Grade", "Nos"]
        _write_filter_row(ws_txn, filters, len(txn_headers))
        _write_headers(ws_txn, 2, txn_headers)
        for i, row in enumerate(data.transactions, 3):
            ws_txn.cell(i, 1, str(row.purchase_date))
            ws_txn.cell(i, 2, row.supplier_name)
            ws_txn.cell(i, 3, row.place)
            ws_txn.cell(i, 4, row.design_name)
            ws_txn.cell(i, 5, row.size)
            ws_txn.cell(i, 6, row.grade_code)
            ws_txn.cell(i, 7, row.nos)

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf
