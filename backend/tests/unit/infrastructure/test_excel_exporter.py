"""V2 TC-199 — ExcelExporter.export_inward_report produces a workbook with exactly
the two expected sheets: Consolidation and Transactions."""
from __future__ import annotations

from datetime import date

from openpyxl import load_workbook

from src.infrastructure.exporters.excel_exporter import ExcelExporter
from src.presentation.schemas.inward_report import (
    InwardConsolidationRow,
    InwardReportResponse,
    InwardTransactionRow,
)


def test_tc199_excel_exporter_inward_report_has_consolidation_and_transactions_sheets():
    data = InwardReportResponse(
        consolidation=[
            InwardConsolidationRow(
                design_id=1,
                design_name="16X10 Ridges",
                size="16X10",
                grade_id=1,
                grade_code="1",
                total_nos=50,
            )
        ],
        transactions=[
            InwardTransactionRow(
                purchase_date=date(2026, 7, 1),
                supplier_id=1,
                supplier_name="Manjunatha",
                place="Mallur",
                design_id=1,
                design_name="16X10 Ridges",
                size="16X10",
                grade_id=1,
                grade_code="1",
                nos=50,
            )
        ],
    )
    buf = ExcelExporter().export_inward_report(data, {})
    buf.seek(0)
    wb = load_workbook(buf)
    assert wb.sheetnames == ["Consolidation", "Transactions"]
    assert len(wb.sheetnames) == 2
