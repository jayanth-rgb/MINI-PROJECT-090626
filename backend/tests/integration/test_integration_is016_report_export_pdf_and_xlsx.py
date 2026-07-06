"""IS-016 — DF-Export: seeded inward+sales → GET /reports/{sales,inward}/export.

Verifies both formats stream correctly for both report types using:
  • Content-Type header check
  • Content-Disposition filename check
  • Magic-byte prefix check (PDF: %PDF-, XLSX: PK\\x03\\x04 ZIP header)
  • openpyxl round-trip check on the XLSX response (Consolidation + Transactions sheets)

Also verifies FastAPI Query pattern rejects unknown formats (422, not 400 — the
pattern regex sits at the validation layer, not inside ReportExportService).
"""
from __future__ import annotations

from datetime import date, timedelta
from io import BytesIO

from openpyxl import load_workbook

from src.infrastructure.db.models.master import (
    DealerModel,
    DesignGradeMapModel,
    GradeModel,
    StaffModel,
    SupplierModel,
    TradingDesignModel,
)


def test_is016(client, db_session):
    # ------------------------------------------------------------------ seed master
    supplier = SupplierModel(supplier_id=1, supplier_name="Manjunatha", place="Mallur")
    dealer = DealerModel(dealer_id=1, dealer_name="Raj Hardwares", place="Mysuru")
    staff = StaffModel(staff_id=1, staff_name="Chandran")
    design = TradingDesignModel(design_id=1, size="16X10", design_name="16X10 Ridges")
    grade = GradeModel(grade_id=1, grade_code="1")
    db_session.add_all([supplier, dealer, staff, design, grade])
    db_session.flush()
    db_session.add(DesignGradeMapModel(design_id=1, grade_id=1, is_active=True))
    db_session.flush()

    txn_date = (date.today() - timedelta(days=1)).isoformat()

    # ---------------------------------------------- seed txn data via HTTP
    resp_in = client.post(
        "/api/v1/inward",
        json={
            "purchase_date": txn_date,
            "supplier_id": 1,
            "entered_by_id": 1,
            "lines": [{"design_id": 1, "grade_id": 1, "nos": 25}],
        },
    )
    assert resp_in.status_code == 201, f"inward failed: {resp_in.text}"

    resp_sale = client.post(
        "/api/v1/sales",
        json={
            "sales_date": txn_date,
            "dealer_id": 1,
            "loading_staff_id": 1,
            "verified_by_id": 1,
            "lines": [{"design_id": 1, "grade_id": 1, "nos": 10}],
        },
    )
    assert resp_sale.status_code == 201, f"sale failed: {resp_sale.text}"

    # ---------------------------------------------- STEP 1 — sales PDF
    resp_sales_pdf = client.get(
        "/api/v1/reports/sales/export", params={"format": "pdf"}
    )
    assert resp_sales_pdf.status_code == 200, f"sales PDF failed: {resp_sales_pdf.text}"
    assert resp_sales_pdf.headers["content-type"].startswith("application/pdf")
    assert resp_sales_pdf.content.startswith(b"%PDF-"), (
        f"sales PDF magic missing; head={resp_sales_pdf.content[:8]!r}"
    )
    cd_sales_pdf = resp_sales_pdf.headers.get("content-disposition", "")
    assert "attachment" in cd_sales_pdf
    assert "sales_report_" in cd_sales_pdf
    assert ".pdf" in cd_sales_pdf

    # ---------------------------------------------- STEP 2 — sales XLSX
    resp_sales_xlsx = client.get(
        "/api/v1/reports/sales/export", params={"format": "xlsx"}
    )
    assert resp_sales_xlsx.status_code == 200, (
        f"sales XLSX failed: {resp_sales_xlsx.text}"
    )
    assert resp_sales_xlsx.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert resp_sales_xlsx.content.startswith(b"PK\x03\x04"), (
        f"sales XLSX ZIP magic missing; head={resp_sales_xlsx.content[:8]!r}"
    )
    cd_sales_xlsx = resp_sales_xlsx.headers.get("content-disposition", "")
    assert "sales_report_" in cd_sales_xlsx
    assert ".xlsx" in cd_sales_xlsx

    # ---------------------------------------------- STEP 3 — inward PDF
    resp_in_pdf = client.get(
        "/api/v1/reports/inward/export", params={"format": "pdf"}
    )
    assert resp_in_pdf.status_code == 200, f"inward PDF failed: {resp_in_pdf.text}"
    assert resp_in_pdf.headers["content-type"].startswith("application/pdf")
    assert resp_in_pdf.content.startswith(b"%PDF-")
    cd_in_pdf = resp_in_pdf.headers.get("content-disposition", "")
    assert "inward_report_" in cd_in_pdf
    assert ".pdf" in cd_in_pdf

    # ---------------------------------------------- STEP 4 — inward XLSX + sheet round-trip
    resp_in_xlsx = client.get(
        "/api/v1/reports/inward/export", params={"format": "xlsx"}
    )
    assert resp_in_xlsx.status_code == 200, f"inward XLSX failed: {resp_in_xlsx.text}"
    assert resp_in_xlsx.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert resp_in_xlsx.content.startswith(b"PK\x03\x04")
    cd_in_xlsx = resp_in_xlsx.headers.get("content-disposition", "")
    assert "inward_report_" in cd_in_xlsx
    assert ".xlsx" in cd_in_xlsx

    wb = load_workbook(BytesIO(resp_in_xlsx.content))
    assert "Consolidation" in wb.sheetnames, (
        f"Consolidation sheet missing; got sheets={wb.sheetnames}"
    )
    assert "Transactions" in wb.sheetnames, (
        f"Transactions sheet missing; got sheets={wb.sheetnames}"
    )

    # ---------------------------------------------- STEP 5 — invalid format → 422
    resp_bad = client.get(
        "/api/v1/reports/sales/export", params={"format": "csv"}
    )
    assert resp_bad.status_code == 422, (
        f"invalid format should be 422, got {resp_bad.status_code}: {resp_bad.text}"
    )
